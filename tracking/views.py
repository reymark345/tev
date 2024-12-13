from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from main.models import (AuthUser, TevIncoming, SystemConfiguration,RoleDetails, StaffDetails, TevOutgoing, TevBridge, Charges, RolePermissions, Division)
import json 
from django.core import serializers
from datetime import date as datetime_date
from django.contrib.auth.hashers import make_password
from django.db import IntegrityError, connection
import math
from django.core.serializers import serialize
from django.forms.models import model_to_dict
import requests
from django.db import connections
from datetime import datetime,timedelta
from receive.filters import UserFilter
import datetime as date_time
from django.db.models import Subquery, Max, F, Q, Exists, OuterRef
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.utils import timezone
from django.template.defaultfilters import date
import decimal

def format_datetime(dt):
    return dt.strftime('%B %d, %Y %I:%M %p') if dt else None

@login_required(login_url='login')
@csrf_exempt
def status(request):
    allowed_roles = ["Admin", "Incoming staff", "Validating staff", "Payroll staff", "Certified staff", "Outgoing staff", "Budget staff", "Journal staff", "Approval staff", "End user"] 

    user_id = request.session.get('user_id', 0)

    role_permissions = RolePermissions.objects.filter(user_id=user_id).values('role_id')
    role_details = RoleDetails.objects.filter(id__in=role_permissions).values('role_name')
    role_names = [entry['role_name'] for entry in role_details]
    
    if any(role_name in allowed_roles for role_name in role_names):
        context = {
            'employee_list' : TevIncoming.objects.filter().order_by('first_name'),
            'permissions' : role_names,
            'division': Division.objects.filter(id__in=[1, 13, 14, 17, 18, 19, 20, 21, 22, 23, 24, 26]).order_by('name')
        }

        return render(request, 'tracking/status.html', context)
    else:
        return redirect("travel-history")

@login_required(login_url='login')
def users(request):
    allowed_roles = ["Admin"]    
    user_id = request.session.get('user_id', 0)
    role_permissions = RolePermissions.objects.filter(user_id=user_id).values('role_id')
    role_details = RoleDetails.objects.filter(id__in=role_permissions).values('role_name')
    role_names = [entry['role_name'] for entry in role_details]

    if any(role_name in allowed_roles for role_name in role_names):
        context = {
            'users' : AuthUser.objects.filter().exclude(id=1).order_by('first_name').select_related(),
            'permissions' : role_names,
            'role_details': RoleDetails.objects.filter().order_by('role_name'),
        }
        return render(request, 'admin/users.html', context)
    else:
        return redirect("travel-history")

        
@login_required(login_url='login')
def status_load(request):
    total = 0
    data = []
    _search = request.GET.get('search[value]')

    FIdNumber= request.GET.get('FIdNumber')
    FStatus= request.GET.get('FStatus') 
    FTransactionCode = request.GET.get('FTransactionCode')
    FDateTravel= request.GET.get('FDateTravel') 
    FDivision = request.GET.get('FDivision')
    DpYear= request.GET.get('DpYear') 

    if DpYear == "2023":
        finance_database_alias = 'finance' 

    elif DpYear == "2025":
        finance_database_alias = 'finance_2025'
    else:
        finance_database_alias = 'finance_2024' 


    year = int(DpYear) % 100
    formatted_year = str(year)+"-"
    year_ = str(year)

    print("testttttyear")
    print(formatted_year)


    
    EmployeeList = request.GET.getlist('EmployeeList[]')
    FAdvancedFilter =  request.GET.get('FAdvancedFilter')

    search_fields = ['code', 'first_name', 'last_name', 'dv_no'] 
    filter_conditions = Q()

    for field in search_fields:
        filter_conditions |= Q(**{f'{field}__icontains': _search})

    if FAdvancedFilter:
        print("conditioneadvancefilter")
        FStartDate = request.GET.get('FStartDate') 
        FEndDate = request.GET.get('FEndDate') 
        formatted_start_date = None
        formatted_end_date = None
        with connection.cursor() as cursor:

            query = """
                SELECT tev_incoming.id, tev_incoming.code, tev_incoming.first_name, tev_incoming.middle_name,
                    tev_incoming.last_name, tev_incoming.date_travel, tev_incoming.status_id,
                    tev_incoming.original_amount, tev_incoming.final_amount, tev_incoming.incoming_in,
                    tev_incoming.incoming_out, tev_incoming.slashed_out,tev_incoming.updated_at, tev_incoming.division, tev_incoming.section,
                    tev_incoming.date_reviewed, tev_incoming.date_payrolled, tev_incoming.review_date_forwarded,
                    tev_bridge.purpose AS purposes, tev_outgoing.dv_no AS dv_no,
                    tev_outgoing.box_date_out, tev_outgoing.box_b_in, tev_outgoing.box_b_out, tev_outgoing.ard_in, tev_outgoing.otg_d_received, tev_outgoing.otg_d_forwarded,
                    tev_outgoing.b_d_received,tev_outgoing.b_d_forwarded, tev_outgoing.j_d_received, tev_outgoing.j_d_forwarded, tev_outgoing.a_d_received, tev_outgoing.a_d_forwarded
                FROM tev_incoming
                INNER JOIN (
                    SELECT MAX(id) AS max_id
                    FROM tev_incoming
                    GROUP BY code
                ) AS latest_ids
                ON tev_incoming.id = latest_ids.max_id
                LEFT JOIN tev_bridge
                ON tev_incoming.id = tev_bridge.tev_incoming_id
                LEFT JOIN tev_outgoing
                ON tev_bridge.tev_outgoing_id = tev_outgoing.id
                WHERE tev_incoming.date_travel LIKE %s
            """
            params = [f'%{DpYear}%']

            # WHERE (tev_outgoing.dv_no LIKE %s OR tev_outgoing.dv_no IS NULL)

            if FTransactionCode:
                query += " AND tev_incoming.code = %s"
                params.append(FTransactionCode)

            if FDateTravel:
                query += " AND tev_incoming.date_travel LIKE %s"
                params.append(f'%{FDateTravel}%')

            if FDivision:
                query += " AND tev_incoming.division = %s"
                params.append(FDivision)

            if FIdNumber:
                query += " AND tev_incoming.id_no = %s"
                params.append(FIdNumber)

            if FStatus:
                if FStatus == "5":
                    FStatus = (4,5,6)
                    query += " AND tev_incoming.status_id IN %s"
                    params.append(FStatus)

                elif FStatus == "8" or FStatus == "9":
                    FStatus = (6,8,9)
                    query += " AND tev_incoming.status_id IN %s"
                    params.append(FStatus)

                elif FStatus == "10" or FStatus == "11":
                    FStatus = (9,10,11)
                    query += " AND tev_incoming.status_id IN %s"
                    params.append(FStatus)

                elif FStatus == "12" or FStatus == "13":
                    FStatus = (11,12,13)
                    query += " AND tev_incoming.status_id IN %s"
                    params.append(FStatus)

                elif FStatus == "14":
                    FStatus = (13,14,15)
                    query += " AND tev_incoming.status_id IN %s"
                    params.append(FStatus)

                else:
                    query += " AND tev_incoming.status_id = %s"
                    params.append(FStatus)

            if EmployeeList:
                placeholders = ', '.join(['%s' for _ in range(len(EmployeeList))])
                query += f" AND tev_incoming.id_no IN ({placeholders})"
                params.extend(EmployeeList)
            query += "ORDER BY tev_incoming.id DESC;"

            cursor.execute(query, params)
            columns = [col[0] for col in cursor.description]
            finance_data = [dict(zip(columns, row)) for row in cursor.fetchall()]

            if FStartDate and FEndDate:
                formatted_start_date = datetime.strptime(FStartDate, '%m/%d/%Y').date()
                formatted_end_date = datetime.strptime(FEndDate, '%m/%d/%Y').date()
                filtered_rows = []

                for row in finance_data:
                    dates = row['date_travel'].split(',')
                    date_objects = [datetime.strptime(date.strip(), '%d-%m-%Y').date() for date in dates if date.strip()]
                    for date in date_objects:
                        if formatted_start_date <= date <= formatted_end_date:
                            filtered_rows.append(row)
                            break
                finance_data = []
                finance_data = filtered_rows

    elif _search:
        print("conditionelifstatement")
        with connection.cursor() as cursor:
            query = """
                SELECT
                    tev_incoming.id,
                    tev_incoming.code,
                    tev_incoming.first_name,
                    tev_incoming.middle_name,
                    tev_incoming.last_name,
                    tev_incoming.date_travel,
                    tev_incoming.status_id,
                    tev_incoming.original_amount,
                    tev_incoming.final_amount,
                    tev_incoming.incoming_in,
                    tev_incoming.incoming_out,
                    tev_incoming.slashed_out,
                    tev_incoming.updated_at,
                    tev_bridge.purpose AS purposes,
                    tev_outgoing.dv_no AS dv_no,
                    tev_incoming.division,
                    tev_incoming.section,
                    tev_incoming.date_reviewed,
                    tev_incoming.date_payrolled,
                    tev_incoming.review_date_forwarded,
                    tev_outgoing.box_date_out, tev_outgoing.box_b_in, tev_outgoing.box_b_out, tev_outgoing.ard_in, tev_outgoing.otg_d_received, tev_outgoing.otg_d_forwarded,
                    tev_outgoing.b_d_received,tev_outgoing.b_d_forwarded, tev_outgoing.j_d_received, tev_outgoing.j_d_forwarded, tev_outgoing.a_d_received, tev_outgoing.a_d_forwarded
                FROM
                    tev_incoming
                INNER JOIN (
                    SELECT
                        MAX(id) AS max_id
                    FROM
                        tev_incoming
                    GROUP BY
                        code
                ) AS latest_ids ON tev_incoming.id = latest_ids.max_id
                LEFT JOIN
                    tev_bridge ON tev_incoming.id = tev_bridge.tev_incoming_id
                LEFT JOIN
                    tev_outgoing ON tev_bridge.tev_outgoing_id = tev_outgoing.id
                WHERE
                    (tev_incoming.code LIKE %s
                    OR tev_incoming.first_name LIKE %s
                    OR tev_incoming.last_name LIKE %s
                    OR tev_incoming.division LIKE %s
                    OR tev_incoming.section LIKE %s
                    OR tev_outgoing.dv_no LIKE %s)
                    AND tev_incoming.date_travel LIKE %s
                ORDER BY
                    tev_incoming.id DESC;
            """
            #AND (tev_outgoing.dv_no LIKE %s OR dv_no IS NULL)
            # cursor.execute(query, [f'%{_search}%', f'%{_search}%', f'%{_search}%', f'%{_search}%', f'%{_search}%', f'%{_search}%', f'%{formatted_year}%'])
            cursor.execute(query, [f'%{_search}%', f'%{_search}%', f'%{_search}%', f'%{_search}%', f'%{_search}%', f'%{_search}%', f'%{DpYear}%'])
            columns = [col[0] for col in cursor.description]
            finance_data = [dict(zip(columns, row)) for row in cursor.fetchall()]
    else:
        print("conditionelsestatement")
        print(formatted_year)
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT tev_incoming.id, tev_incoming.code, tev_incoming.first_name, tev_incoming.middle_name,
                    tev_incoming.last_name, tev_incoming.date_travel, tev_incoming.status_id,
                    tev_incoming.original_amount, tev_incoming.final_amount, tev_incoming.incoming_in,
                    tev_incoming.incoming_out,tev_incoming.slashed_out,tev_incoming.updated_at, tev_incoming.division, tev_incoming.section,
                    tev_incoming.date_reviewed, tev_incoming.date_payrolled, tev_incoming.review_date_forwarded,
                    tev_bridge.purpose AS purposes, tev_outgoing.dv_no AS dv_no,
                    tev_outgoing.box_date_out, tev_outgoing.box_b_in, tev_outgoing.box_b_out, tev_outgoing.ard_in, tev_outgoing.otg_d_received, tev_outgoing.otg_d_forwarded,
                    tev_outgoing.b_d_received,tev_outgoing.b_d_forwarded, tev_outgoing.j_d_received, tev_outgoing.j_d_forwarded, tev_outgoing.a_d_received, tev_outgoing.a_d_forwarded
                FROM tev_incoming
                INNER JOIN (
                    SELECT MAX(id) AS max_id
                    FROM tev_incoming
                    GROUP BY code
                ) AS latest_ids
                ON tev_incoming.id = latest_ids.max_id
                LEFT JOIN tev_bridge
                ON tev_incoming.id = tev_bridge.tev_incoming_id
                LEFT JOIN tev_outgoing
                ON tev_bridge.tev_outgoing_id = tev_outgoing.id
                WHERE tev_incoming.date_travel LIKE %s
                ORDER BY tev_incoming.id DESC;
            """, [f'%{DpYear}%'])
            print("DpYear")
            print(year_)

            # WHERE (tev_outgoing.dv_no LIKE %s OR tev_outgoing.dv_no IS NULL)
            #     ORDER BY tev_incoming.id DESC;
            # """, [f'{formatted_year}%'])
            
            columns = [col[0] for col in cursor.description]
            finance_data = [dict(zip(columns, row)) for row in cursor.fetchall()]
    
    total = len(finance_data)
    _start = request.GET.get('start')
    _length = request.GET.get('length')
    if _start and _length:
        start = int(_start)
        length = int(_length)
        page = math.ceil(start / length) + 1
        per_page = length
        finance_data = finance_data[start:start + length]

    for row in finance_data:
        amt_certified = ''
        amt_journal = ''
        amt_budget = ''
        amt_check = ''
        approved_date = ''

        finance_query = """
            SELECT ts.dv_no, ts.amt_certified, ts.amt_journal, ts.amt_budget, tc.check_amount, ts.approval_date
            FROM transactions AS ts
            LEFT JOIN trans_check AS tc ON tc.dv_no = ts.dv_no WHERE ts.dv_no = %s
        """
        if row['dv_no']:
            with connections[finance_database_alias].cursor() as cursor2:
                cursor2.execute(finance_query, (row['dv_no'],))
                finance_results = cursor2.fetchall()

            if finance_results:
                amt_certified = finance_results[0][1]
                amt_journal = finance_results[0][2]
                amt_budget = finance_results[0][3]
                amt_check = finance_results[0][4]
                approved_date = finance_results[0][5]
                
        first_name = row['first_name'] if row['first_name'] else ''
        middle_name = row['middle_name'] if row['middle_name'] else ''
        last_name = row['last_name'] if row['last_name'] else ''
        
        emp_fullname = f"{first_name} {middle_name} {last_name}".strip()

        acronym = row['division']
        section = row['section']

        acr = ("4Ps" if acronym == 'Pantawid Pamilyang Pilipino Program' else 
          "AD" if acronym == 'Administrative Division' else
          "FMD" if acronym == 'Financial Management Division' else
          "DRMD" if acronym == 'Disaster Response Management Division' else
          "HRMDD" if acronym == 'Human Resource Management and Development Division' else
          "PSD" if acronym == 'Protective Services Division' else
          "PPD" if acronym == 'Policy and Plans Division' else
          "ORD" if acronym == 'Office of the Regional Director' else
          "AD" if acronym == 'Administrative Division' else
          "PD/SLP" if acronym == 'Promotive Services Division' and section == 'Sustainable Livelihood Program'  else
          "PD" if acronym == 'Promotive Services Division'else
          "")
        
        if approved_date:
            approved_date = approved_date.strftime('%B %d, %Y')
        else:
            approved_date = None
        
        item = {
            'division': acr,
            'code': row['code'],
            'full_name': emp_fullname,
            'date_travel': row['date_travel'],
            'status': row['status_id'],
            'original_amount': row['original_amount'],
            'final_amount': row['final_amount'],
            'incoming_in': format_datetime(row['incoming_in']) if row['incoming_in'] is not None else '',
            'incoming_out': format_datetime(row['incoming_out']) if row['incoming_out'] is not None else '',
            'slashed_out': format_datetime(row['slashed_out']) if row['slashed_out'] is not None else '',
            'updated_at': format_datetime(row['updated_at']) if row['updated_at'] is not None else '',
            'date_reviewed': format_datetime(row['date_reviewed']) if row['date_reviewed'] is not None else '',
            'date_payrolled': format_datetime(row['date_payrolled']) if row['date_payrolled'] is not None else '',
            'review_date_forwarded': format_datetime(row['review_date_forwarded']) if row['review_date_forwarded'] is not None else '',
            'box_date_out': format_datetime(row['box_date_out']) if row['box_date_out'] is not None else '',
            'box_b_in': format_datetime(row['box_b_in']) if row['box_b_in'] is not None else '',
            'box_b_out': format_datetime(row['box_b_out']) if row['box_b_out'] is not None else '',
            'ard_in': format_datetime(row['ard_in']) if row['ard_in'] is not None else '',
            'otg_d_received': format_datetime(row['otg_d_received']) if row['otg_d_received'] is not None else '',
            'otg_d_forwarded': format_datetime(row['otg_d_forwarded']) if row['otg_d_forwarded'] is not None else '',
            'b_d_received': format_datetime(row['b_d_received']) if row['b_d_received'] is not None else '',
            'b_d_forwarded': format_datetime(row['b_d_forwarded']) if row['b_d_forwarded'] is not None else '',
            'j_d_received': format_datetime(row['j_d_received']) if row['j_d_received'] is not None else '',
            'j_d_forwarded': format_datetime(row['j_d_forwarded']) if row['j_d_forwarded'] is not None else '',
            'a_d_received': format_datetime(row['a_d_received']) if row['a_d_received'] is not None else '',
            'a_d_forwarded': format_datetime(row['a_d_forwarded']) if row['a_d_forwarded'] is not None else '',
            'purpose': row['purposes'],
            'dv_no': row['dv_no'],
            'id': row['id'],
            'amt_certified': amt_certified,
            'amt_journal': amt_journal,
            'amt_budget': amt_budget,
            'amt_check': amt_check,
            'approved_date': approved_date,
            'section': row['section']
        }
        data.append(item)
        
    response = {
        'data': data,
        'page': page,
        'per_page': per_page,
        'recordsTotal': total,
        'recordsFiltered': total,
    }
    return JsonResponse(response)


@login_required(login_url='login')
@csrf_exempt
def employee_details(request):
    year= request.POST.get('DpYear')
    finance_database_alias = 'finance_2024' if year == "2024" else 'finance' 
    allowed_roles = ["Admin", "Incoming staff", "Validating staff"] 
    fullname = ''
    total_amount = 0
    charges_list = []
    data = []  
    idd = request.POST.get('dv_id')
    incoming = TevIncoming.objects.filter(id=idd).first()
    amt_certified = ''
    amt_journal = ''
    amt_budget = ''
    amt_check = ''
    ap_date = ''
    check_issued_status = ''
    check_issued_date = ''
    check_issued_released = ''

    query = """ 
        SELECT 
                ti.id, 
                ti.code,
                ti.id_no,
                ti.first_name,
                ti.middle_name,
                ti.last_name,
                ti.original_amount, 
                ti.final_amount, 
                ti.status_id, 
                GROUP_CONCAT(tb.purpose SEPARATOR ', ') AS purposes,
                ti.incoming_in,
                ti.incoming_out,
                CONCAT_WS(' ', ui.first_name, ui.last_name) AS incoming_by,
                ti.date_payrolled,
                CONCAT_WS(' ', pb.first_name, pb.last_name) AS payrolled_by,
                t_o.dv_no, 
                t_o.box_b_out,
                t_o.otg_d_received,
                CONCAT_WS(' ', ot_r.first_name, ot_r.last_name) AS otg_r_user_id,
                t_o.otg_d_forwarded,
                CONCAT_WS(' ', ot_f.first_name, ot_f.last_name) AS otg_out_user_id,
                t_o.b_d_received,
				CONCAT_WS(' ', b_r.first_name, b_r.last_name) AS b_r_user_id,
                t_o.b_d_forwarded,
                CONCAT_WS(' ', b_f.first_name, b_f.last_name) AS b_out_user_id,
                t_o.j_d_received,
				CONCAT_WS(' ', j_r.first_name, j_r.last_name) AS j_r_user_id,
                t_o.j_d_forwarded,
                CONCAT_WS(' ', j_f.first_name, j_f.last_name) AS j_out_user_id,
                t_o.a_d_received,
				CONCAT_WS(' ', a_r.first_name, a_r.last_name) AS a_r_user_id,
                t_o.a_d_forwarded,
                CONCAT_WS(' ', a_f.first_name, a_f.last_name) AS a_out_user_id,
                CONCAT_WS(' ', ob.first_name, ob.last_name) AS out_by,
                ch.name AS charges, 
                cl.name AS cluster,
                GROUP_CONCAT(
                    CONCAT(
                        '<strong><u>', t3.name, '</u></strong>', 
                        ' - ', 
                        DATE_FORMAT(t2.date, '%%M %%d, %%Y')
                    ) SEPARATOR '; '
                ) AS formatted_remarks,
                CONCAT_WS(' ', au.first_name, au.last_name) AS forwarded_by,
                CONCAT_WS(' ', rb.first_name, rb.last_name) AS reviewed_by,
                ti.date_reviewed,
                ti.review_date_forwarded,
                CONCAT_WS(' ', rf.first_name, rf.last_name) AS review_forwarded_by
        FROM 
                tev_incoming AS ti 
        LEFT JOIN 
                tev_bridge AS tb ON tb.tev_incoming_id = ti.id
        LEFT JOIN 
                tev_outgoing AS t_o ON t_o.id = tb.tev_outgoing_id
        LEFT JOIN 
                charges AS ch ON ch.id = tb.charges_id
        LEFT JOIN 
                cluster AS cl ON cl.id = t_o.cluster
        LEFT JOIN 
				auth_user AS ob ON ob.id = t_o.out_by
        LEFT JOIN 
                remarks_r AS t2 ON t2.incoming_id = ti.id
        LEFT JOIN 
                remarks_lib AS t3 ON t3.id = t2.remarks_lib_id
        LEFT JOIN 
				auth_user AS au ON au.id = ti.forwarded_by
        LEFT JOIN 
				auth_user AS rb ON rb.id = ti.reviewed_by
        LEFT JOIN 
				auth_user AS ui ON ui.id = ti.user_id
        LEFT JOIN 
				auth_user AS pb ON pb.id = ti.payrolled_by
        LEFT JOIN 
				auth_user AS rf ON rf.id = ti.review_forwarded_by
        LEFT JOIN 
				auth_user AS ot_r ON ot_r.id = t_o.otg_r_user_id
        LEFT JOIN 
				auth_user AS ot_f ON ot_f.id = t_o.otg_out_user_id
        LEFT JOIN 
                auth_user AS b_r ON b_r.id = t_o.b_r_user_id
        LEFT JOIN 
                auth_user AS b_f ON b_f.id = t_o.b_out_user_id
        LEFT JOIN 
                auth_user AS j_r ON j_r.id = t_o.j_r_user_id
        LEFT JOIN 
                auth_user AS j_f ON j_f.id = t_o.j_out_user_id
        LEFT JOIN 
                auth_user AS a_r ON a_r.id = t_o.a_r_user_id
        LEFT JOIN 
                auth_user AS a_f ON a_f.id = t_o.a_out_user_id
        WHERE 
                ti.code LIKE %s
        GROUP BY 
                ti.id, 
                ti.code, 
                ti.id_no,
                ti.original_amount, 
                ti.final_amount, 
                ti.status_id, 
                ti.remarks, 
                ti.incoming_in,
                ti.incoming_out,
                ti.date_payrolled,
                ti.payrolled_by,
                t_o.dv_no,
                t_o.box_b_out,
                t_o.otg_d_received,
                t_o.otg_r_user_id,
                t_o.otg_d_forwarded, 
                t_o.otg_out_user_id,
                t_o.b_d_received,
				t_o.b_r_user_id,
                t_o.b_d_forwarded,
                t_o.b_out_user_id,
                t_o.j_d_received,
				t_o.j_r_user_id,
                t_o.j_d_forwarded,
                t_o.j_out_user_id,
                t_o.a_d_received,
				t_o.a_r_user_id,
                t_o.a_d_forwarded,
                t_o.a_out_user_id,
                t_o.out_by,
                ch.name, 
                cl.name
        ORDER BY 
                ti.id DESC;
    """

    with connection.cursor() as cursor:
        cursor.execute(query, [incoming.code])
        results = cursor.fetchall()


    id_number = incoming.id_no

    def format_date(date_str):
        if date_str is not None:
            if isinstance(date_str, datetime):
                return date_str.strftime("%B %d, %Y %I:%M %p")
            else:
                return datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S.%f").strftime("%B %d, %Y %I:%M %p")
        else:
            return ''
        
    for row in results:
        incoming_in = row[10]
        incoming_out = row[11]
        incoming = row[12]
        p_date = row[13]
        box_b_out = row[16]
        otg_d_received = row[17]
        otg_r_by = row[18]
        otg_d_f = row[19]
        otg_f_user_id = row[20]
        budget_d_r = row[21]
        budget_r_by = row[22]
        budget = row[23]
        budget_f_by = row[24]
        journal_d_r = row[25]
        journal_r_by = row[26]
        journal = row[27]
        journal_f_by = row[28]
        approval_d_r = row[29]
        approval_r_by = row[30]
        approval = row[31]
        approval_f_by = row[32]
        forwarded_by = row[37]
        date_reviewed = date(row[39], "F j, Y g:i A")
        review_date_forwarded = row[40]
        review_forwarded_by = row[41]

        if row[15]:
            finance_query = """
                SELECT ts.dv_no, ts.amt_certified, ts.amt_journal, ts.amt_budget, tc.check_amount, ts.approval_date, tp.check_id, tp.check_issued, tp.check_released
                FROM transactions AS ts
                LEFT JOIN trans_check AS tc ON tc.dv_no = ts.dv_no
                LEFT JOIN trans_payeename AS tp ON tp.dv_no = ts.dv_no  WHERE ts.dv_no = %s
            """
            with connections[finance_database_alias].cursor() as cursor2:
                cursor2.execute(finance_query, (row[15],))
                finance_results = cursor2.fetchall()

            if finance_results:
                amt_certified = finance_results[0][1]
                amt_journal = finance_results[0][2]
                amt_budget = finance_results[0][3]
                amt_check = finance_results[0][4]
                approved_date = finance_results[0][5]
                check_issued_status = finance_results[0][6]
                check_issued_date = finance_results[0][7]
                check_issued_released = finance_results[0][8]
                if approved_date:
                    approved_date = datetime.strptime(str(approved_date), "%Y-%m-%d")
                    formatted_date = approved_date.strftime("%B %d, %Y")
                    ap_date = f"{formatted_date}"
                if check_issued_date:
                    check_issued_date = datetime.strptime(str(check_issued_date), "%Y-%m-%d")
                    formatted_date = check_issued_date.strftime("%B %d, %Y")
                    check_issued_date = f"{formatted_date}"
                if check_issued_released:
                    check_issued_released = datetime.strptime(str(check_issued_released), "%Y-%m-%d")
                    formatted_date = check_issued_released.strftime("%B %d, %Y")
                    check_issued_released = f"{formatted_date}"

        item = {
            'id': row[0],
            'code': row[1],
            'id_no': row[2],
            'emp_fullname': row[3] + " "+ row[4] + " " + row[5],
            'original_amount': row[6],
            'final_amount': row[7],
            'status': row[8],
            'purpose': row[9], 
            'incoming_in': format_date(incoming_in),
            'incoming_out': format_date(incoming_out),
            'incoming_by': incoming.title(),
            'p_d': format_date(p_date),
            'p_by': row[14],
            'dv_no': row[15],
            'p_d_f': format_date(box_b_out),
            'otg_d_received': format_date(otg_d_received),
            'otg_r_by': otg_r_by.title(),
            'otg_d_f': format_date(otg_d_f),
            'otg_f_user': otg_f_user_id.title(),
            'b_d_r': format_date(budget_d_r),
            'b_r_by': budget_r_by.title(),
            'b_d_f': format_date(budget),
            'b_f_by': budget_f_by.title(),
            'j_d_r': format_date(journal_d_r),
            'j_r_by': journal_r_by.title(),
            'j_d_f': format_date(journal),
            'j_f_by': journal_f_by.title(),
            'a_d_received': format_date(approval_d_r),
            'a_r_by': approval_r_by.title(),
            'a_d_f': format_date(approval),
            'a_f_by': approval_f_by.title(),
            'p_f_by': row[33],
            'charges': row[34],
            'cluster': row[35],
            "remarks": row[36],
            'received_forwarded_by': forwarded_by.title(), 
            'reviewed_by': row[38],
            'date_reviewed': date_reviewed,
            'review_date_forwarded': format_date(review_date_forwarded) if review_date_forwarded else '',
            'review_forwarded_by': review_forwarded_by if review_forwarded_by else '',
            'amt_certified': amt_certified,
            'amt_journal': amt_journal,
            'amt_budget': amt_budget,
            'amt_check': amt_check,
            'certified_date': ap_date,
            'check_issued_status': check_issued_status,
            'check_issued_date' : check_issued_date,
            'check_issued_released' : check_issued_released,
        }
        data.append(item)   

    total = len(data)    

          
    response = {
        'data': data,
        'full_name': fullname,
        'id_number': id_number,
        'charges': charges_list,
        'is_print': 1,
        'recordsTotal': total,
        'recordsFiltered': total,
        'total_amount':total_amount
    }
    return JsonResponse(response)

@login_required(login_url='login')
@csrf_exempt
def travel_history(request):
    user_id = request.session.get('user_id', 0)
    role_permissions = RolePermissions.objects.filter(user_id=user_id).values('role_id')
    role_details = RoleDetails.objects.filter(id__in=role_permissions).values('role_name')
    role_names = [entry['role_name'] for entry in role_details]
    context = {
        'employee_list' : TevIncoming.objects.filter().order_by('first_name'),
        'permissions' : role_names,
    }
    return render(request, 'tracking/travel_history.html', context)

@csrf_exempt
def export_status(request):
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT tev_incoming.id,tev_outgoing.dv_no AS dv_no, tev_incoming.code, tev_incoming.account_no, tev_incoming.id_no,tev_incoming.last_name, tev_incoming.first_name, tev_incoming.middle_name,
                tev_incoming.date_travel, tev_incoming.division, tev_incoming.section,charges.name, st.name, au.first_name AS incoming_by,rb.first_name AS reviewed_by,
                tev_incoming.original_amount, tev_incoming.final_amount, payrolled_charges.amount, tev_incoming.incoming_in AS date_actual, tev_incoming.updated_at AS date_entry, tev_incoming.date_reviewed,
                tev_incoming.review_date_forwarded AS date_reviewed_forwarded, tev_bridge.purpose AS purposes, pb.first_name AS payrolled_by, tev_incoming.date_payrolled
            FROM tev_incoming
            INNER JOIN (
            SELECT MAX(id) AS max_id
            FROM tev_incoming
            GROUP BY code
            ) AS latest_ids
            ON tev_incoming.id = latest_ids.max_id
            LEFT JOIN tev_bridge
            ON tev_incoming.id = tev_bridge.tev_incoming_id
            LEFT JOIN tev_outgoing
            ON tev_bridge.tev_outgoing_id = tev_outgoing.id
            LEFT JOIN payrolled_charges
            ON payrolled_charges.incoming_id = tev_incoming.id
            LEFT JOIN charges
            ON payrolled_charges.charges_id = charges.id
            LEFT JOIN auth_user AS au
            ON au.id = tev_incoming.user_id
            LEFT JOIN auth_user AS rb
            ON rb.id = tev_incoming.reviewed_by
            LEFT JOIN auth_user AS pb
            ON pb.id = tev_incoming.payrolled_by
            LEFT JOIN status AS st
            ON st.id = tev_incoming.status_id
            ORDER BY tev_incoming.id DESC;
        """)
        rows = cursor.fetchall()

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename={date}-TRIS-REPORT.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'REPORT'

    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )
    columns = [
        'DV NO',
        'CODE',
        'ACCOUNT NO',
        'ID NO',
        'LASTNAME',
        'FIRSTNAME',
        'MIDDLE INITIAL',
        'DATE TRAVEL',
        'DIVISION',
        'SECTION',
        'CHARGES',
        'STATUS',
        'INCOMING BY',
        'REVIEWED BY',
        'ORIGINAL AMOUNT',
        'AMOUT CERTIFIED',
        'CHARGES AMOUNT',
        'DATE ACTUAL RECEIVED',
        'DATE ENTRY',
        'DATE REVIEWED',
        'DATE REVIEWED FORWARDED',
        'PURPOSE',
        'PAYROLLED BY',
        'PAYROLLED DATE',

    ]
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]


    for tris in rows:
        original_amount = round(float(tris[15]), 2) if tris[15] is not None else None
        final_amount = round(float(tris[16]), 2) if tris[16] is not None else None
        charges_amount = round(float(tris[17]), 2) if tris[17] is not None else 0

        if original_amount is not None:
            original_amount = '{:.2f}'.format(original_amount)

        if final_amount is not None:
            final_amount = '{:.2f}'.format(final_amount)

        if charges_amount is not None:
            charges_amount = '{:.2f}'.format(charges_amount)

        row_num += 1
        row = [
            tris[1],  # dv_no
            tris[2],  # code
            tris[3],  # account_no
            tris[4],  # id_no
            tris[5],  # last_name
            tris[6],  # middle_name
            tris[7],  # first_name
            tris[8],  # date_travel
            tris[9],  # division
            tris[10], # section
            tris[11],  # charges
            tris[12],  # status_id
            tris[13],  # incoming_by
            tris[14],  # reviewed_by
            original_amount,
            final_amount,
            charges_amount,
            tris[18],
            tris[19],
            tris[20],  
            tris[21],  
            tris[22],  
            tris[23], 
            tris[24], 
        ]       
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = 30
    workbook.save(response)
    return response

@login_required(login_url='login')
@csrf_exempt
def travel_calendar(request):
    allowed_roles = ["Admin", "Incoming staff", "Validating staff","Payroll staff","Certified staff", "End user"] 
    user_id = request.session.get('user_id', 0)
    role_permissions = RolePermissions.objects.filter(user_id=user_id).values('role_id')
    role_details = RoleDetails.objects.filter(id__in=role_permissions).values('role_name')
    role_names = [entry['role_name'] for entry in role_details]


    if any(role_name in allowed_roles for role_name in role_names):
        context = {
            'employee_list' : TevIncoming.objects.filter().order_by('first_name'),
            'permissions' : role_names,
        }
        return render(request, 'tracking/travel_calendar.html', context)
    else:
        return redirect("travel-history")
        # return render(request, 'pages/unauthorized.html')

def travel_history_load(request):
    total = 0
    data = []
    _search = request.GET.get('search[value]')
    user_id = request.session.get('user_id', 0)
    user_details_ = AuthUser.objects.filter(id=user_id).first()
    first_name_ = user_details_.first_name
    last_name_ = user_details_.last_name
    FStatus= request.GET.get('FStatus') 
    FTransactionCode = request.GET.get('FTransactionCode')
    FDateTravel= request.GET.get('FDateTravel') 
    FDivision = request.GET.get('FDivision')
    DpYear= request.GET.get('DpYear') 

    if DpYear == "2023":
        finance_database_alias = 'finance' 
    else:
        finance_database_alias = 'finance_2024' 

    year = int(DpYear) % 100
    formatted_year = str(year)+"-"
    
    EmployeeList = request.GET.getlist('EmployeeList[]')
    FAdvancedFilter =  request.GET.get('FAdvancedFilter')

    search_fields = ['code', 'first_name', 'last_name', 'dv_no'] 
    filter_conditions = Q()

    for field in search_fields:
        filter_conditions |= Q(**{f'{field}__icontains': _search})

    if FAdvancedFilter:
        FStartDate = request.GET.get('FStartDate') 
        FEndDate = request.GET.get('FEndDate') 
        formatted_start_date = None
        formatted_end_date = None
        with connection.cursor() as cursor:
            query = """
                SELECT tev_incoming.id, tev_incoming.code, tev_incoming.first_name, tev_incoming.middle_name,
                    tev_incoming.last_name, tev_incoming.date_travel, tev_incoming.status_id,
                    tev_incoming.original_amount, tev_incoming.final_amount, tev_incoming.incoming_in,
                    tev_incoming.incoming_out, tev_incoming.division, tev_incoming.section,
                    tev_bridge.purpose AS purposes, tev_outgoing.dv_no AS dv_no
                FROM tev_incoming
                INNER JOIN (
                    SELECT MAX(id) AS max_id
                    FROM tev_incoming
                    GROUP BY code
                ) AS latest_ids
                ON tev_incoming.id = latest_ids.max_id
                LEFT JOIN tev_bridge
                ON tev_incoming.id = tev_bridge.tev_incoming_id
                LEFT JOIN tev_outgoing
                ON tev_bridge.tev_outgoing_id = tev_outgoing.id
                WHERE (tev_outgoing.dv_no LIKE %s OR tev_outgoing.dv_no IS NULL) AND tev_incoming.first_name = %s and last_name = %s
            """
            params = [f'{formatted_year}%', first_name_, last_name_]

            if FTransactionCode:
                query += " AND tev_incoming.code = %s"
                params.append(FTransactionCode)

            if FDateTravel:
                query += " AND tev_incoming.date_travel LIKE %s"
                params.append(f'%{FDateTravel}%')

            if FDivision:
                query += " AND tev_incoming.division = %s"
                params.append(FDivision)

            if FStatus:
                if FStatus == "5":
                    FStatus = (4,5,6)
                    query += " AND tev_incoming.status_id IN %s"
                    params.append(FStatus)

                elif FStatus == "8" or FStatus == "9":
                    FStatus = (6,8,9)
                    query += " AND tev_incoming.status_id IN %s"
                    params.append(FStatus)

                elif FStatus == "10" or FStatus == "11":
                    FStatus = (9,10,11)
                    query += " AND tev_incoming.status_id IN %s"
                    params.append(FStatus)

                elif FStatus == "12" or FStatus == "13":
                    FStatus = (11,12,13)
                    query += " AND tev_incoming.status_id IN %s"
                    params.append(FStatus)

                elif FStatus == "14":
                    FStatus = (13,14,15)
                    query += " AND tev_incoming.status_id IN %s"
                    params.append(FStatus)

                else:
                    query += " AND tev_incoming.status_id = %s"
                    params.append(FStatus)

            if EmployeeList:
                placeholders = ', '.join(['%s' for _ in range(len(EmployeeList))])
                query += f" AND tev_incoming.id_no IN ({placeholders})"
                params.extend(EmployeeList)
            query += "ORDER BY tev_incoming.id DESC;"

            cursor.execute(query, params)
            columns = [col[0] for col in cursor.description]
            finance_data = [dict(zip(columns, row)) for row in cursor.fetchall()]

            if FStartDate and FEndDate:
                formatted_start_date = datetime.strptime(FStartDate, '%m/%d/%Y').date()
                formatted_end_date = datetime.strptime(FEndDate, '%m/%d/%Y').date()
                filtered_rows = []

                for row in finance_data:
                    dates = row['date_travel'].split(',')
                    date_objects = [datetime.strptime(date.strip(), '%d-%m-%Y').date() for date in dates if date.strip()]
                    for date in date_objects:
                        if formatted_start_date <= date <= formatted_end_date:
                            filtered_rows.append(row)
                            break
                finance_data = []
                finance_data = filtered_rows

    elif _search:
        with connection.cursor() as cursor:
            query = """
                SELECT
                    tev_incoming.id,
                    tev_incoming.code,
                    tev_incoming.first_name,
                    tev_incoming.middle_name,
                    tev_incoming.last_name,
                    tev_incoming.date_travel,
                    tev_incoming.status_id,
                    tev_incoming.original_amount,
                    tev_incoming.final_amount,
                    tev_incoming.incoming_in,
                    tev_incoming.incoming_out,
                    tev_bridge.purpose AS purposes,
                    tev_outgoing.dv_no AS dv_no,
                    tev_incoming.division,
                    tev_incoming.section
                FROM
                    tev_incoming
                INNER JOIN (
                    SELECT
                        MAX(id) AS max_id
                    FROM
                        tev_incoming
                    GROUP BY
                        code
                ) AS latest_ids ON tev_incoming.id = latest_ids.max_id
                LEFT JOIN
                    tev_bridge ON tev_incoming.id = tev_bridge.tev_incoming_id
                LEFT JOIN
                    tev_outgoing ON tev_bridge.tev_outgoing_id = tev_outgoing.id
                WHERE
                    (tev_incoming.code LIKE %s
                    OR tev_incoming.first_name LIKE %s
                    OR tev_incoming.last_name LIKE %s
                    OR tev_incoming.division LIKE %s
                    OR tev_incoming.section LIKE %s
                    OR tev_outgoing.dv_no LIKE %s)
                    AND (tev_outgoing.dv_no LIKE %s OR dv_no IS NULL) AND tev_incoming.first_name = %s and last_name = %s
                ORDER BY
                    tev_incoming.id DESC;
            """
            cursor.execute(query, [f'%{_search}%', f'%{_search}%', f'%{_search}%', f'%{_search}%', f'%{_search}%', f'%{_search}%', f'%{formatted_year}%', f'{first_name_}', f'{last_name_}'])
            columns = [col[0] for col in cursor.description]
            finance_data = [dict(zip(columns, row)) for row in cursor.fetchall()]
    else:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT tev_incoming.id, tev_incoming.code, tev_incoming.first_name, tev_incoming.middle_name,
                    tev_incoming.last_name, tev_incoming.date_travel, tev_incoming.status_id,
                    tev_incoming.original_amount, tev_incoming.final_amount, tev_incoming.incoming_in,
                    tev_incoming.incoming_out, tev_incoming.division, tev_incoming.section,
                    tev_bridge.purpose AS purposes, tev_outgoing.dv_no AS dv_no
                FROM tev_incoming
                INNER JOIN (
                    SELECT MAX(id) AS max_id
                    FROM tev_incoming
                    GROUP BY code
                ) AS latest_ids
                ON tev_incoming.id = latest_ids.max_id
                LEFT JOIN tev_bridge
                ON tev_incoming.id = tev_bridge.tev_incoming_id
                LEFT JOIN tev_outgoing
                ON tev_bridge.tev_outgoing_id = tev_outgoing.id
                WHERE (tev_outgoing.dv_no LIKE %s OR tev_outgoing.dv_no IS NULL) AND tev_incoming.first_name = %s and last_name = %s
                ORDER BY tev_incoming.id DESC;
            """, [f'{formatted_year}%', first_name_, last_name_])


            columns = [col[0] for col in cursor.description]
            finance_data = [dict(zip(columns, row)) for row in cursor.fetchall()]
    
    total = len(finance_data)
    _start = request.GET.get('start')
    _length = request.GET.get('length')
    if _start and _length:
        start = int(_start)
        length = int(_length)
        page = math.ceil(start / length) + 1
        per_page = length
        finance_data = finance_data[start:start + length]

    for row in finance_data:
        amt_certified = ''
        amt_journal = ''
        amt_budget = ''
        amt_check = ''
        approved_date = ''

        finance_query = """
            SELECT ts.dv_no, ts.amt_certified, ts.amt_journal, ts.amt_budget, tc.check_amount, ts.approval_date
            FROM transactions AS ts
            LEFT JOIN trans_check AS tc ON tc.dv_no = ts.dv_no WHERE ts.dv_no = %s
        """
        if row['dv_no']:
            with connections[finance_database_alias].cursor() as cursor2:
                cursor2.execute(finance_query, (row['dv_no'],))
                finance_results = cursor2.fetchall()

            if finance_results:
                amt_certified = finance_results[0][1]
                amt_journal = finance_results[0][2]
                amt_budget = finance_results[0][3]
                amt_check = finance_results[0][4]
                approved_date = finance_results[0][5]
                
        first_name = row['first_name'] if row['first_name'] else ''
        middle_name = row['middle_name'] if row['middle_name'] else ''
        last_name = row['last_name'] if row['last_name'] else ''
        
        emp_fullname = f"{first_name} {middle_name} {last_name}".strip()

        acronym = row['division']
        section = row['section']

        acr = ("4Ps" if acronym == 'Pantawid Pamilyang Pilipino Program' else 
          "AD" if acronym == 'Administrative Division' else
          "FMD" if acronym == 'Financial Management Division' else
          "DRMD" if acronym == 'Disaster Response Management Division' else
          "HRMDD" if acronym == 'Human Resource Management and Development Division' else
          "PSD" if acronym == 'Protective Services Division' else
          "PPD" if acronym == 'Policy and Plans Division' else
          "ORD" if acronym == 'Office of the Regional Director' else
          "AD" if acronym == 'Administrative Division' else
          "PD/SLP" if acronym == 'Promotive Services Division' and section == 'Sustainable Livelihood Program'  else
          "PD" if acronym == 'Promotive Services Division'else
          "")
        
        item = {
            'division': acr,
            'code': row['code'],
            'full_name': emp_fullname,
            'date_travel': row['date_travel'],
            'status': row['status_id'],
            'original_amount': row['original_amount'],
            'final_amount': row['final_amount'],
            'incoming_in': row['incoming_in'],
            'incoming_out': row['incoming_out'],
            'purpose': row['purposes'],
            'dv_no': row['dv_no'],
            'id': row['id'],
            'amt_certified': amt_certified,
            'amt_journal': amt_journal,
            'amt_budget': amt_budget,
            'amt_check': amt_check,
            'approved_date': approved_date,
            'section': row['section']
        }
        data.append(item)
        
    response = {
        'data': data,
        'page': page,
        'per_page': per_page,
        'recordsTotal': total,
        'recordsFiltered': total,
    }
    return JsonResponse(response)








