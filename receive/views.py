from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from main.models import (AuthUser, TevIncoming, SystemConfiguration,RoleDetails, StaffDetails, TevOutgoing, TevBridge, RemarksLib, Remarks_r, RolePermissions, Division, Section, TransactionLogs, TravelList, TravelDestination )
import json 
from django.core import serializers
from datetime import date as datetime_date
from django.contrib.auth.hashers import make_password
from django.db import IntegrityError, connection
import math
from django.core.serializers import serialize
from django.forms.models import model_to_dict
import requests
from django.db.models import Q, F, Exists, OuterRef
from django.db import connections
from datetime import datetime,timedelta
from receive.filters import UserFilter
import datetime as date_time
from openpyxl import load_workbook
from tablib import Dataset
import ast
from django.db.models import F, CharField, Value
from django.db.models.functions import Concat
from django.utils import timezone
from django.template.defaultfilters import date
from decimal import Decimal
from suds.client import Client
from django.db import transaction
from django.conf import settings
import platform
import re
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def generate_code():
    trans_code = SystemConfiguration.objects.values_list(
        'transaction_code', flat=True
    ).first()

    last_code = trans_code.split('-')
    sample_date = datetime_date.today()
    year = sample_date.strftime("%y")
    month = sample_date.strftime("%m")
    day = sample_date.strftime("%d")
    if last_code[0] == year:
        series = int(last_code[2]) + 1
    else:
        series = 1
    code = year + '-' + month + '-' + f'{series:05d}'
    return code



@login_required(login_url='login')
def list(request):
    allowed_roles = ["Admin", "Incoming staff", "Validating staff"] 
    user_id = request.session.get('user_id', 0)
    role_permissions = RolePermissions.objects.filter(user_id=user_id).values('role_id')
    role_details = RoleDetails.objects.filter(id__in=role_permissions).values('role_name')
    role_names = [entry['role_name'] for entry in role_details]
    data = []
    user_name = RolePermissions.objects.filter(role_id=2) 
    get_id = user_name.values_list('user_id', flat=True)
    date_actual = SystemConfiguration.objects.filter().first().date_actual
    for user_id in get_id:
        userData = AuthUser.objects.filter(id=user_id)
        full_name = userData[0].first_name + ' ' + userData[0].last_name if userData else ''
        item_entry = {
            'id': userData[0].id,
            'full_name': full_name
        }
        data.append(item_entry)
    if any(role_name in allowed_roles for role_name in role_names):
        context = {
            'employee_list' : TevIncoming.objects.filter().order_by('first_name'),
            'remarks_list' : RemarksLib.objects.filter(status = 1).order_by('name'),
            'is_actual_date': date_actual,
            'permissions' : role_names,
            'created_by' :  data
        }
        return render(request, 'receive/receive.html' , context)
    else:
        return render(request, 'pages/unauthorized.html')
    
@login_required(login_url='login')
def travel_list(request):
    
    allowed_roles = ["Admin", "Incoming staff", "Validating staff"] 
    user_id = request.session.get('user_id', 0)
    role_permissions = RolePermissions.objects.filter(user_id=user_id).values('role_id')
    role_details = RoleDetails.objects.filter(id__in=role_permissions).values('role_name')
    role_names = [entry['role_name'] for entry in role_details]
    data = []
    user_name = RolePermissions.objects.filter(role_id=2) 
    get_id = user_name.values_list('user_id', flat=True)
    date_actual = SystemConfiguration.objects.filter().first().date_actual
    for user_id in get_id:
        userData = AuthUser.objects.filter(id=user_id)
        full_name = userData[0].first_name + ' ' + userData[0].last_name if userData else ''
        item_entry = {
            'id': userData[0].id,
            'full_name': full_name
        }
        data.append(item_entry)
    if any(role_name in allowed_roles for role_name in role_names):
        context = {
            'employee_list' : TevIncoming.objects.filter().order_by('first_name'),
            'remarks_list' : RemarksLib.objects.filter(status = 1).order_by('name'),
            'is_actual_date': date_actual,
            'permissions' : role_names,
            'created_by' :  data
        }
        return render(request, 'receive/travel_user.html' , context)
    else:
        return render(request, 'pages/unauthorized.html')
    
@csrf_exempt
def api(request):
    url = settings.PORTAL_API_URL
    portal_token = settings.PORTAL_TOKEN
    headers = {
        "Authorization": portal_token,
    }
    try:
        response = requests.get(url, headers=headers, verify=False)
        data = response.json()
        return JsonResponse({'data': data})
    except Exception as e:
        return JsonResponse({'error': 'An error occurred', 'details': str(e)}, status=500)

# @csrf_exempt
# def api(request):
#     url = settings.PORTAL_API_URL
#     portal_token = settings.PORTAL_TOKEN
#     headers = {
#         "Authorization": portal_token,
#     }
#     response = requests.get(url, headers=headers)
#     data = response.json()
#     return JsonResponse({'data': data})

@csrf_exempt
def psgc_api(request):
    province_url = settings.PSGC_PROVINCE_URL
    access_token = settings.PSGC_TOKEN
    headers = {"Authorization": f"Bearer {access_token}"}

    try:
        prov_response = requests.get(province_url, headers=headers)
        prov_response.raise_for_status()
        prov_api_data = prov_response.json()
        province_list = []

        for province in prov_api_data['data']['provinces']:
            province_data = {
                "prov_id": province["prov_id"],
                "prov_code_correspondence": province["prov_code_correspondence"],
                "prov_name": province["prov_name"],
                "prov_code": province["prov_code"],
                "geo_level": province["geo_level"],
                "old_name": province.get("old_name"),
                "income_classification": province["income_classification"],
                "region_code": province["region_code"],
                "region_code_correspondence": province["region_code_correspondence"],
                "reg_id": province["reg_id"],
                "cities": []
            }
            city_url = f"{settings.PSGC_CITY_URL}{province['prov_code']}" 
            
            city_response = requests.get(city_url, headers=headers)
            city_response.raise_for_status()
            city_api_data = city_response.json()

            for city in city_api_data['data']['municipalities']:
                city_data = {
                    "city_id": city["city_id"],
                    "city_code_correspondence": city["city_code_correspondence"],
                    "city_name": city["city_name"],
                    "city_code": city["city_code"],
                    "classification": city["classification"],
                    "old_name": city.get("old_name"),
                    "city_class": city["city_class"],
                    "income_classification": city["income_classification"],
                    "province_code": city["province_code"],
                    "province_code_correspondence": city["province_code_correspondence"],
                    "prov_id": city["prov_id"],
                    "barangays": []
                }
                # barangay_url = f"https://dxcloud.dswd.gov.ph/api/psgc/barangayByMunicipality?municipality={city['city_code']}"
                barangay_url = f"{settings.PSGC_BARANGAY_URL}{city['city_code']}" 
                brgy_response = requests.get(barangay_url, headers=headers)
                brgy_response.raise_for_status()
                brgy_api_data = brgy_response.json()

                for barangay in brgy_api_data['data']['barangay']:
                    barangay_data = {
                        "brgy_id": barangay["brgy_id"],
                        "brgy_code_correspondence": barangay["brgy_code_correspondence"],
                        "brgy_name": barangay["brgy_name"],
                        "brgy_code": barangay["brgy_code"],
                        "geo_level": barangay["geo_level"],
                        "old_name": barangay.get("old_name"),
                        "city_class": barangay.get("city_class"),
                        "urb_rur": barangay["urb_rur"],
                        "city_code": barangay["city_code"],
                        "city_code_correspondence": barangay["city_code_correspondence"],
                        "city_id": barangay["city_id"]
                    }
                    city_data["barangays"].append(barangay_data)
                province_data["cities"].append(city_data)
            province_list.append(province_data)
        return JsonResponse(province_list, safe=False)

    except requests.exceptions.RequestException as e:
        return JsonResponse({"error": str(e)}, status=500)
    
@login_required(login_url='login')
def checking(request):
    allowed_roles = ["Admin", "Incoming staff", "Validating staff"] 
    user_id = request.session.get('user_id', 0)

    role_permissions = RolePermissions.objects.filter(user_id=user_id).values('role_id')
    role_details = RoleDetails.objects.filter(id__in=role_permissions).values('role_name')
    role_names = [entry['role_name'] for entry in role_details]
    date_actual = SystemConfiguration.objects.filter().first().date_actual

    created_user_name = RolePermissions.objects.filter(role_id=2)
    user_name = RolePermissions.objects.filter(role_id=3) 
    created_get_id = created_user_name.values_list('user_id', flat=True)
    get_id = user_name.values_list('user_id', flat=True)
    data = []
    created_data = []

    for c_user_id in created_get_id:
        c_userData = AuthUser.objects.filter(id=c_user_id)
        c_full_name = c_userData[0].first_name + ' ' + c_userData[0].last_name if c_userData else ''
        c_item_entry = {
            'id': c_userData[0].id,
            'full_name': c_full_name
        }
        created_data.append(c_item_entry)
    
    for user_id in get_id:
        userData = AuthUser.objects.filter(id=user_id)
        full_name = userData[0].first_name + ' ' + userData[0].last_name if userData else ''
        item_entry = {
            'id': userData[0].id,
            'full_name': full_name
        }
        data.append(item_entry)


    if any(role_name in allowed_roles for role_name in role_names):
        context = {
            'employee_list' : TevIncoming.objects.filter().order_by('first_name'),
            'remarks_list' : RemarksLib.objects.filter(status = 1).order_by('name'),
            'permissions' : role_names,
            'is_actual_date': date_actual,
            'created_by' :  created_data,
            'reviewed_by' :  data
        }
        return render(request, 'receive/review_docs.html', context)
    else:
        return render(request, 'pages/unauthorized.html')
    
    
@login_required(login_url='login')
@csrf_exempt
def search_list(request):
    allowed_roles = ["Admin", "Incoming staff", "Validating staff"] 
    return JsonResponse({'data': "success"})


def item_load(request):
    _search = request.GET.get('search[value]', '').strip()
    FAdvancedFilter = request.GET.get('FAdvancedFilter')
    EmployeeList = request.GET.getlist('EmployeeList[]')
    id_numbers = EmployeeList if EmployeeList else []
    year = request.GET.get('DpYear')
    year = int(year)

    base_query = """
        SELECT t1.*, 
               GROUP_CONCAT(CONCAT('<strong><u>', t3.name, '</u></strong> - ', DATE_FORMAT(t2.date, '%%M %%d, %%Y')) SEPARATOR '<br>') AS formatted_remarks,
               GROUP_CONCAT(t3.name SEPARATOR ', ') AS lacking
        FROM tev_incoming t1
        LEFT JOIN remarks_r t2 ON t2.incoming_id = t1.id
        LEFT JOIN remarks_lib t3 ON t3.id = t2.remarks_lib_id
        WHERE (t1.code, t1.id) IN (
            SELECT DISTINCT code, MAX(id)
            FROM tev_incoming
            GROUP BY code
        )
        AND ((`status_id` IN (3) AND slashed_out IS NOT NULL) 
             OR (`status_id` IN (1) AND slashed_out IS NULL))
    """
    params = []
    if FAdvancedFilter:
        advanced_filters = """
            {id_no_filter}
            AND account_no LIKE %s
            AND date_travel LIKE %s
            AND original_amount LIKE %s
            AND final_amount LIKE %s
            AND incoming_in LIKE %s
            AND user_id LIKE %s
            AND status_id LIKE %s
        """
        if id_numbers:
            id_no_filter = "AND id_no IN %s"
            params.append(tuple(id_numbers))
        else:
            id_no_filter = ""
        params.extend([
            '%' + request.GET.get('FAccountNumber', '') + '%',
            '%' + request.GET.get('FDateTravel', '') + '%',
            '%' + request.GET.get('FOriginalAmount', '') + '%',
            '%' + request.GET.get('FFinalAmount', '') + '%',
            '%' + request.GET.get('FIncomingIn', '') + '%',
            '%' + request.GET.get('FCreatedBy', '') + '%',
            '%' + request.GET.get('FStatus', '') + '%'
        ])

        base_query += advanced_filters.format(id_no_filter=id_no_filter)
    elif _search:
        search_filters = """
            AND (first_name LIKE %s OR last_name LIKE %s 
                 OR id_no LIKE %s OR original_amount LIKE %s OR final_amount LIKE %s)
        """
        base_query += search_filters
        params.extend(['%' + _search + '%'] * 5)

    date_travel_filter = " AND date_travel LIKE %s"
    base_query += date_travel_filter
    params.append(f"%{year}%")
    base_query += " GROUP BY t1.id ORDER BY id DESC;"
    
    with connection.cursor() as cursor:
        cursor.execute(base_query, params)
        columns = [col[0] for col in cursor.description]
        results = [dict(zip(columns, row)) for row in cursor.fetchall()]

    total = len(results)
    _start = int(request.GET.get('start', 0))
    _length = int(request.GET.get('length', 10))
    results = results[_start:_start + _length]

    data = []
    for item in results:
        userData = AuthUser.objects.filter(id=item['user_id']).first()
        full_name = userData.first_name if userData else ''
        
        emp_fullname = f"{item['first_name']} {item['middle_name']} {item['last_name']}".strip()

        item_entry = {
            'id': item['id'],
            'code': item['code'],
            'name': emp_fullname,
            'id_no': item['id_no'],
            'account_no': item['account_no'],
            'date_travel': item['date_travel'],
            'original_amount': item['original_amount'],
            'final_amount': item['final_amount'],
            'incoming_in': item['incoming_in'],
            'incoming_out': item['incoming_out'],
            'slashed_out': item['slashed_out'],
            'remarks': item['remarks'],
            'lacking': item['formatted_remarks'],
            'status': item['status_id'],
            'user_id': full_name
        }

        data.append(item_entry)

    response = {
        'data': data,
        'recordsTotal': total,
        'recordsFiltered': total,
    }
    return JsonResponse(response)

def travel_loadss(request):    
    data = []
    user_data = TransactionLogs.objects.all().order_by('-id')

    total = len(user_data)
    page = 1
    per_page = total

    _start = request.GET.get('start')
    _length = request.GET.get('length')
    
    if _start and _length:
        start = int(_start)
        length = int(_length)
        page = math.ceil(start / length) + 1
        per_page = length
        user_data = user_data[start:start + length]
        
    for item in user_data:
        userData = AuthUser.objects.filter(id=item.user_id)
        full_name = userData[0].first_name + ' ' + userData[0].last_name
       
        user_data_item = {
            'description': item.description,
            'user': full_name.upper(),
            'created_at': item.created_at,
        }
        data.append(user_data_item)

    response = {
        'data': data,
        'page': page,
        'per_page': per_page,
        'recordsTotal': total,
        'recordsFiltered': total,
    }
    return JsonResponse(response)


def travel_load(request):
    user_id = request.session.get('user_id', 0)
    results = TravelList.objects.all().order_by('-id')


    total = len(results)
    _start = request.GET.get('start')
    _length = request.GET.get('length')
    if _start and _length:
        start = int(_start)
        length = int(_length)
        page = math.ceil(start / length) + 1
        per_page = length
        results = results[start:start + length]

    data = []


    for item in results:

        item_entry = {
            'id': item.id,
            'code': item.date,
            'name': item.date,
            'id_no': item.date,
            'account_no': item.date,
            'date_travel': item.date,
            'original_amount': item.date,
            'final_amount': item.date,
            'incoming_in': item.date,
            'incoming_out': item.date,
            'slashed_out':item.date,
            'remarks': item.date,
            'lacking':item.date,
            'status': item.date,
            'user_id': 'fasfasfasf'
        }

        data.append(item_entry)

    response = {
        'data': data,
        'page': page,
        'per_page': per_page,
        'recordsTotal': total,
        'recordsFiltered': total,
    }
    return JsonResponse(response)


def checking_load(request):
    _search = request.GET.get('search[value]')
    FIdNumber= request.GET.get('FIdNumber')
    FTransactionCode = request.GET.get('FTransactionCode')
    FDateTravel= request.GET.get('FDateTravel') 
    FIncomingIn= request.GET.get('FIncomingIn')
    FOriginalAmount= request.GET.get('FOriginalAmount')
    FFinalAmount= request.GET.get('FFinalAmount')
    FAccountNumber= request.GET.get('FAccountNumber')
    FAdvancedFilter =  request.GET.get('FAdvancedFilter')
    FStatus = request.GET.get('FStatus')
    EmployeeList = request.GET.getlist('EmployeeList[]')
    FReviewedBy = request.GET.get('FReviewedBy')
    FCreatedBy = request.GET.get('FCreatedBy')
    year = request.GET.get('DpYear')

    status_txt = ''
    if _search in "returned":
        status_txt = '3'

    elif _search in "for checking":
        status_txt = '2'

    elif _search in "pending":
        status_txt = '16'
    else:
        status_txt = '7'
        
    id_numbers = EmployeeList if EmployeeList else []

    if FAdvancedFilter:
        query = """
            SELECT t1.*, GROUP_CONCAT(t3.name SEPARATOR ', ') AS lacking
            FROM tev_incoming t1
            LEFT JOIN remarks_r AS t2 ON t2.incoming_id = t1.id
            LEFT JOIN remarks_lib AS t3 ON t3.id = t2.remarks_lib_id
            WHERE (t1.status_id = 2
                OR t1.status_id = 7
                OR t1.status_id = 16
                OR t1.status_id = 17
                OR (t1.status_id = 3 AND t1.slashed_out IS NULL))
        """
        params = []

        if FStatus:
            query += " AND t1.status_id = %s"
            params.append(FStatus)

        if FTransactionCode:
            query += " AND t1.code = %s"
            params.append(FTransactionCode)

        if FIdNumber:
            query += " AND t1.id_no = %s"
            params.append(FIdNumber)

        if FDateTravel:
            query += " AND t1.date_travel LIKE %s"
            params.append(f'%{FDateTravel}%')

        if FIncomingIn:
            query += " AND t1.incoming_in = %s"
            params.append(FIncomingIn)

        if FAccountNumber:
            query += " AND t1.account_no = %s"
            params.append(FAccountNumber)

        if FOriginalAmount:
            query += " AND t1.original_amount = %s"
            params.append(FOriginalAmount)

        if FFinalAmount:
            query += " AND t1.final_amount = %s"
            params.append(FFinalAmount)
        
        if FCreatedBy:
            query += " AND t1.user_id = %s"
            params.append(FCreatedBy)

        if FReviewedBy:
            query += " AND t1.reviewed_by = %s"
            params.append(FReviewedBy)

        if EmployeeList:
            placeholders = ', '.join(['%s' for _ in range(len(EmployeeList))])
            query += f" AND t1.id_no IN ({placeholders})"
            params.extend(EmployeeList)

        query += " AND date_travel LIKE %s"
        params.append(f"%{year}%")
        query += " GROUP BY t1.id ORDER BY t1.incoming_out DESC;"


    elif _search:
        query = """
            SELECT t1.*, GROUP_CONCAT(t3.name SEPARATOR ', ') AS lacking
            FROM tev_incoming t1
            LEFT JOIN remarks_r AS t2 ON t2.incoming_id = t1.id
            LEFT JOIN remarks_lib AS t3 ON t3.id = t2.remarks_lib_id
            WHERE (t1.status_id = 2
                            OR t1.status_id = 7
                            OR t1.status_id = 16
                            OR t1.status_id = 17
                            OR (t1.status_id = 3 AND t1.slashed_out IS NULL)
            )
            AND (code LIKE %s
            OR first_name LIKE %s
            OR last_name LIKE %s
            OR id_no LIKE %s
            OR original_amount LIKE %s
            OR final_amount LIKE %s
            )
            AND date_travel LIKE %s
            GROUP BY t1.id ORDER BY id DESC;
        """

        params = [
            '%' + _search + '%' if _search else "%%",
            '%' + _search + '%' if _search else "%%",
            '%' + _search + '%' if _search else "%%",
            '%' + _search + '%' if _search else "%%",
            '%' + _search + '%' if _search else "%%",
            '%' + _search + '%' if _search else "%%",
            '%' + year + '%' if year else "%%",
        ]
    else:

        query = """
            SELECT t1.*, GROUP_CONCAT(t3.name SEPARATOR ', ') AS lacking
            FROM tev_incoming t1
            LEFT JOIN remarks_r AS t2 ON t2.incoming_id = t1.id
            LEFT JOIN remarks_lib AS t3 ON t3.id = t2.remarks_lib_id
            WHERE (t1.status_id = 2
                    OR t1.status_id = 7
                    OR t1.status_id = 16
                    OR t1.status_id = 17
                    OR (t1.status_id = 3 AND t1.slashed_out IS NULL)
            )
            AND (code LIKE %s
            OR id_no LIKE %s
            OR account_no LIKE %s
            OR date_travel LIKE %s
            OR original_amount LIKE %s
            OR final_amount LIKE %s
            OR remarks LIKE %s
            OR status_id LIKE %s
            )
            AND date_travel LIKE %s
            GROUP BY t1.id ORDER BY id DESC;
        """
        params = ['%' + _search + '%', '%' + _search + '%', '%' + _search + '%', '%' + _search + '%', '%' + _search + '%', '%' + _search + '%', '%' + _search + '%','%' + status_txt + '%','%' + year + '%']
    
    with connection.cursor() as cursor:
        cursor.execute(query, params)
        columns = [col[0] for col in cursor.description]
        results = [dict(zip(columns, row)) for row in cursor.fetchall()]

    total = len(results)
    _start = request.GET.get('start')
    _length = request.GET.get('length')
    if _start and _length:
        start = int(_start)
        length = int(_length)
        page = math.ceil(start / length) + 1
        per_page = length
        results = results[start:start + length]

    data = []

    for item in results:
        userData = AuthUser.objects.filter(id=item['user_id'])
        full_name = userData[0].first_name if userData else ''
        
        first_name = item['first_name'] if item['first_name'] else ''
        middle_name = item['middle_name'] if item['middle_name'] else ''
        last_name = item['last_name'] if item['last_name'] else ''
        
        emp_fullname = f"{first_name} {middle_name} {last_name}".strip()
        formatted_date_out = date(item['incoming_out'], "F j, Y g:i A")

        item_entry = {
            'id': item['id'],
            'code': item['code'],
            'name': emp_fullname,
            'id_no': item['id_no'],
            'account_no': item['account_no'],
            'date_travel': item['date_travel'],
            'original_amount': item['original_amount'],
            'final_amount': item['final_amount'],
            'incoming_in': item['incoming_in'],
            'incoming_out': formatted_date_out,
            'remarks': item['remarks'],
            'lacking': item['lacking'],
            'status': item['status_id'],
            'user_id': full_name
        }

        data.append(item_entry)

    response = {
        'data': data,
        'page': page,
        'per_page': per_page,
        'recordsTotal': total,
        'recordsFiltered': total,
    }
    return JsonResponse(response)



def read_excel_file(excel_file):
    workbook = load_workbook(excel_file, data_only=True)
    worksheet = workbook.active
    has_empty_fields = False

    excel_data = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        id_no, amount, date_travel = row
        
        if amount is None:
            amount = 0
        
        dates = []
        if date_travel:
            date_list = date_travel.split(',')
            for date_str in date_list:
                date_obj = datetime.strptime(date_str.strip(), '%d-%m-%Y')
                dates.append(date_obj.strftime('%d-%m-%Y')) 

        excel_data.append({
            'id_no': id_no,
            'amount': amount,
            'date_travel': dates,
        })

    for record in excel_data:
        if record['id_no'] is None or record['amount'] == 0 or not record['date_travel']:
            has_empty_fields = True
            break

    if has_empty_fields:
        return has_empty_fields
    else:
        return excel_data

@csrf_exempt
def upload_tev(request):
    user_id = request.session.get('user_id', 0)

    if request.method == 'POST' and request.FILES.get('ExcelData'):
        excel_file = request.FILES['ExcelData']
        if not excel_file.name.endswith('xlsx'):
            return JsonResponse({'data': 'errorxlsx'})

        try:
            sc_code = SystemConfiguration.objects.first()
            sc_code = sc_code.transaction_code
            matched_data = []
            id_list = []
            duplicate_travel = []
            duplicate_te = []

            formatted_duplicates = []
            duplicate_dates = {}
            seen_dates = {}

            employees_data = json.loads(request.POST.get('employees'))
            excel_data = read_excel_file(excel_file)

            if excel_data ==True:
                response_data = {
                    'data': 'empty'
                }
                return JsonResponse(response_data) 

            for row in excel_data:
                g_code = generate_code()
                id_no, amount, date_travel = row['id_no'], row['amount'], row['date_travel']
                

                for date in date_travel:
                    if id_no not in seen_dates:
                        seen_dates[id_no] = set()
                    if date in seen_dates[id_no]:
                        if id_no not in duplicate_dates:
                            duplicate_dates[id_no] = []
                        duplicate_dates[id_no].append(date)
                    seen_dates[id_no].add(date)
                
                id_number_value = None
                formatted_date_travel = ', '.join(date_travel).replace(', ', ',')
                for employee in employees_data:
                    
                    if employee.get('idNumber') == id_no:
                        id_number_value = employee.get('idNumber')
                        acc_no_value = employee.get('accNumber')
                        first_name_value = employee.get('firstName')
                        middle_initial_value = employee.get('middleInitial')
                        last_name_value = employee.get('lastName')

                dates_to_check = formatted_date_travel.split(',')
                duplicate_records = {}

                for date in dates_to_check:
                    results = TevIncoming.objects.filter(date_travel__contains=date).filter(id_no=id_number_value).all()
                    if results:
                        for record in results:
                            if date not in duplicate_records:
                                duplicate_records[date] = []
                            duplicate_records[date].append({'id_no': record.id_no, 'date_travel': record.date_travel})


                for date, records in duplicate_records.items():
                    for record in records:
                        duplicate_te.append({
                            'id_no': record["id_no"],
                            'travel': date
                        })
                        
                if results:
                    # duplicate_travel.append(formatted_date_travel)
                    duplicate_travel.append({
                        'id_no': id_no,
                        'duplicate_travel': formatted_date_travel
                    })

                if id_number_value:
                    matched_data.append({
                        'id_no': id_no,
                        'g_code': g_code,
                        'amount': amount,
                        'date_travel': formatted_date_travel,
                        'idNumber': id_number_value,
                        'accNumber': acc_no_value,
                        'firstName': first_name_value,
                        'middleInitial': middle_initial_value,
                        'lastName': last_name_value
                    })
                    system_config = SystemConfiguration.objects.first()
                    system_config.transaction_code = g_code
                    system_config.save()

                else:
                    id_list.append(id_no)
            if id_list:
                system_config = SystemConfiguration.objects.first()
                system_config.transaction_code = sc_code
                system_config.save()
                response_data = {
                    'data': 'success',
                    'id_no': id_list
                }
                return JsonResponse(response_data) 
            
            elif duplicate_dates:
                for id_no, dates in duplicate_dates.items():
                    formatted_entry = {
                        'id_no': id_no,
                        'duplicate_travel': ','.join(dates)
                    }
                    formatted_duplicates.append(formatted_entry)
                response_data = {
                    'data': 'success',
                    'duplicate_excel_dates': formatted_duplicates
                }
                return JsonResponse(response_data) 

            elif duplicate_travel:
                response_data = {
                    'data': 'success',
                    'duplicate_travel': duplicate_travel
                }
                return JsonResponse(response_data) 
            else:
                try:
                    matched_data_list = matched_data
                    tev_incoming_instances = [
                        TevIncoming(
                            code=data['g_code'],
                            original_amount=data['amount'],
                            id_no=data['id_no'],
                            account_no=data['accNumber'],
                            date_travel=data['date_travel'],
                            first_name=data['firstName'],
                            middle_name=data['middleInitial'],
                            last_name=data['lastName'],
                            user_id = user_id,
                            is_upload = True
                            )
                            for data in matched_data_list
                        ]
                    
                    TevIncoming.objects.bulk_create(tev_incoming_instances)
                    return JsonResponse({'data': 'success'})
        
            
                except json.JSONDecodeError as e:
                   print(f"Error decoding JSON: {e}")
                   matched_data_list = []

        except json.JSONDecodeError:
            return JsonResponse({'data': 'errorjson'})

    else:
        return JsonResponse({'data': 'error'})

def item_edit(request):
    id = request.GET.get('id')
    items = TevIncoming.objects.get(pk=id)
    data = serialize("json", [items])
    return HttpResponse(data, content_type="application/json")

def preview_received(request):

    id = request.GET.get('id')
    with connection.cursor() as cursor:
        query = """
        SELECT
            t1.id,
            t1.code,
            t1.first_name,
            t1.middle_name,
            t1.last_name,
            t1.id_no,
            t1.account_no,
            t1.date_travel,
            t1.original_amount,
            t1.final_amount,
            t1.incoming_in,
            t1.incoming_out,
            t1.slashed_out,
            t1.remarks,
            t1.user_id,
            t1.status_id,
            GROUP_CONCAT(t3.id SEPARATOR ', ') AS lacking,
            GROUP_CONCAT(t2.date SEPARATOR ', ') AS date_remarks
        FROM
            tev_incoming t1
            LEFT JOIN remarks_r AS t2 ON t2.incoming_id = t1.id
            LEFT JOIN remarks_lib AS t3 ON t3.id = t2.remarks_lib_id
        WHERE
            t1.id = %s
        """
        cursor.execute(query, [id])
        result = cursor.fetchone()

    if result:
        orig_amt =  Decimal(result[8])
        orig_amt = orig_amt.quantize(Decimal("0.01"))
        data = {
            'id': result[0],
            'code': result[1],
            'first_name': result[2],
            'middle_name': result[3],
            'last_name': result[4],
            'id_no': result[5],
            'account_no': result[6],
            'date_travel': result[7],
            'original_amount': orig_amt,
            'final_amount': result[9],
            'incoming_in': result[10],
            'incoming_out': result[11],
            'slashed_out': result[12],
            'remarks': result[13],
            'user_id': result[14],
            'status_id': result[15],
            'lacking': result[16],
            'date_remarks': result[17],
        }
        return JsonResponse(data)
    else:
        return JsonResponse({'error': 'No data found for the given ID'}, status=404)

@csrf_exempt
def item_update(request):
    id = request.POST.get('ItemID')
    name = request.POST.get('EmpName')
    middle = request.POST.get('EmpMiddle')
    lname = request.POST.get('EmpLastname')
    amount = request.POST.get('OriginalAmount')
    selected_remarks = request.POST.getlist('selectedRemarks[]')
    selected_dates = request.POST.getlist('selectedDate[]')
    travel_date = request.POST.get('DateTravel')
    range_travel = request.POST.get('RangeTravel')
    date_received = request.POST.get('DateReceived')
    id_no = request.POST.get('IdNumber')
    acc_no = request.POST.get('AccountNumber')
    div = request.POST.get('Division')
    sec = request.POST.get('Section')
    contact = request.POST.get('Contact')

    individual_dates = travel_date.split(',')
    enable_expiry = SystemConfiguration.objects.filter().first().is_travel_expire
    days_expire = SystemConfiguration.objects.filter().first().days_expire
    expired_dates = []


    if travel_date:
        travel_date = request.POST.get('DateTravel')
    else :
        start_date_str, end_date_str = range_travel.split(' to ')
        
        start_date = datetime.strptime(start_date_str.strip(), '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str.strip(), '%Y-%m-%d')
        
        formatted_dates = []

        current_date = start_date
        while current_date <= end_date:
            formatted_dates.append(current_date.strftime('%d-%m-%Y'))
            current_date += timedelta(days=1)

        formatted_dates_str = ', '.join(formatted_dates)
        travel_date = formatted_dates_str

    if enable_expiry:
        for date_str in individual_dates:
            date_object = datetime.strptime(date_str.strip(), '%d-%m-%Y').date()  
            if (datetime_date.today() - date_object).days >= int(days_expire):
                expired_dates.append(date_object.strftime('%B %d, %Y'))

        if expired_dates:
            print("Expired dates found:", expired_dates)
            return JsonResponse({'data': 'expired', 'message': ', '.join(expired_dates)})
        else:
            print("All dates are not expired")

    duplicate_travel = []
    individual_dates = travel_date.split(',')
    cleaned_dates = ','.join(date.strip() for date in individual_dates)
    for date in individual_dates:
        cleaned_date = date.strip()
        results = Q(first_name=name) & \
          Q(middle_name=middle) & \
          Q(last_name=lname) & \
          Q(date_travel__icontains=cleaned_date) & \
          ~Q(status_id=3) & \
          ~Q(id=id)
        
        results = TevIncoming.objects.filter(results).values('date_travel')
        
        if results:
            duplicate_travel.append(cleaned_date)

    if duplicate_travel:
        formatted_dates = [date.replace("'", "") for date in duplicate_travel]
        result = ",".join(formatted_dates)

        date_components = result.split(',')
        def format_date(date_str):
            date_object = datetime.strptime(date_str, '%d-%m-%Y')
            formatted_date = date_object.strftime('%b. %d, %Y')
            return formatted_date
        formatted_dates = [format_date(date) for date in date_components]
        formatted_dates_string = ', '.join(formatted_dates)
        formatted_dates_string = formatted_dates_string
        
        TevIncoming.objects.filter(id=id).update(first_name=name,middle_name = middle,last_name = lname, id_no = id_no, account_no = acc_no, date_travel = travel_date, original_amount=amount, incoming_in = date_received, remarks = formatted_dates_string, division = div, section = sec, contact_no = contact)
        Remarks_r.objects.filter(incoming_id=id).delete()
        for selected_remarks, selected_dates in zip(selected_remarks, selected_dates):
            remarks_lib = Remarks_r(
                date=selected_dates,
                incoming_id=id,
                remarks_lib_id=selected_remarks
            )
            remarks_lib.save()
        
        return JsonResponse({'data': 'error', 'message': duplicate_travel})
    else:
        TevIncoming.objects.filter(id=id).update(first_name=name,middle_name = middle,last_name = lname, id_no = id_no, account_no = acc_no, date_travel = travel_date, original_amount=amount, incoming_in = date_received, remarks = None, division = div, section = sec, contact_no = contact)
        Remarks_r.objects.filter(incoming_id=id).delete()
        for selected_remarks, selected_dates in zip(selected_remarks, selected_dates):
            remarks_lib = Remarks_r(
                date=selected_dates,
                incoming_id=id,
                remarks_lib_id=selected_remarks
            )
            remarks_lib.save()
        return JsonResponse({'data': 'success'})
    

@csrf_exempt
def item_rod_update(request):

    id = request.POST.get('ItemID')
    name = request.POST.get('EmpName')
    middle = request.POST.get('EmpMiddle')
    lname = request.POST.get('EmpLastname')
    travel_date = request.POST.get('DateTravel')
    date_received = request.POST.get('DateReceived')
    id_no = request.POST.get('IdNumber')
    acc_no = request.POST.get('AccountNumber')
    div = request.POST.get('Division')
    sec = request.POST.get('Section')
    orig_amnt = request.POST.get('FAmount')
    duplicate_travel = []
    individual_dates = travel_date.split(',')

    for date in individual_dates:
        cleaned_date = date.strip()
        results = Q(first_name=name) & \
          Q(middle_name=middle) & \
          Q(last_name=lname) & \
          Q(date_travel__icontains=cleaned_date) & \
          ~Q(status_id=3) & \
          ~Q(id=id)
        
        results = TevIncoming.objects.filter(results).values('date_travel')
        
        if results:
            duplicate_travel.append(cleaned_date)

    if duplicate_travel:
        formatted_dates = [date.replace("'", "") for date in duplicate_travel]
        result = ",".join(formatted_dates)

        date_components = result.split(',')
        def format_date(date_str):
            date_object = datetime.strptime(date_str, '%d-%m-%Y')
            formatted_date = date_object.strftime('%b. %d, %Y')
            return formatted_date
        formatted_dates = [format_date(date) for date in date_components]
        formatted_dates_string = ', '.join(formatted_dates)
        formatted_dates_string = formatted_dates_string
        TevIncoming.objects.filter(id=id).update(first_name=name,middle_name = middle,last_name = lname, id_no = id_no, account_no = acc_no,date_travel = travel_date, incoming_in = date_received, remarks = formatted_dates_string, division = div, section = sec, original_amount = orig_amnt)
        return JsonResponse({'data': 'error', 'message': duplicate_travel})
    else:
        TevIncoming.objects.filter(id=id).update(first_name=name,middle_name = middle,last_name = lname, id_no = id_no, account_no = acc_no,date_travel = travel_date, incoming_in = date_received, remarks = None, division = div, section = sec, original_amount = orig_amnt)
        return JsonResponse({'data': 'success'})

@csrf_exempt
def item_returned(request):
    id = request.POST.get('ItemID')
    travel_date = request.POST.get('HDateTravel')
    selected_remarks = request.POST.getlist('selectedRemarks[]')
    selected_dates = request.POST.getlist('selectedDate[]')

    travel_date_stripped = travel_date.strip()
    travel_date_spaces = travel_date_stripped.replace(' ', '')
    id = request.POST.get('ItemID')

    data = TevIncoming.objects.filter(id=id).first()
    tev_add = TevIncoming(code=data.code,first_name=data.first_name,middle_name=data.middle_name,last_name = data.last_name,id_no = data.id_no, account_no = data.account_no, date_travel = travel_date_spaces,original_amount=data.original_amount,final_amount = data.final_amount,incoming_in =date_time.datetime.now(),user_id=data.user_id, division = data.division, section = data.section, contact_no = data.contact_no)
    tev_add.save()

    last_added_tevincoming = TevIncoming.objects.latest('id')
    for selected_remarks, selected_dates in zip(selected_remarks, selected_dates):
        remarks_lib = Remarks_r(
            date=selected_dates,
            incoming_id=last_added_tevincoming.id,
            remarks_lib_id=selected_remarks
        )
        remarks_lib.save()

    return JsonResponse({'data': 'success'})



@csrf_exempt
def item_add(request):
    amount = request.POST.get('OriginalAmount')
    travel_date = request.POST.get('DateTravel')
    range_travel = request.POST.get('RangeTravel')
    date_received = request.POST.get('DateReceived')
    idd_no = request.POST.get('IdNumber')
    acct_no = request.POST.get('AccountNumber')
    contact = request.POST.get('Contact')
    name = request.POST.get('EmpName')
    middle = request.POST.get('EmpMiddle')
    lname = request.POST.get('EmpLastname')
    user_id = request.session.get('user_id', 0)
    division = request.POST.get('Division')
    section = request.POST.get('Section')
    selected_remarks = request.POST.getlist('selectedRemarks[]')
    selected_dates = request.POST.getlist('selectedDate[]')
    g_code = generate_code()
    enable_expiry = SystemConfiguration.objects.filter().first().is_travel_expire
    days_expire = SystemConfiguration.objects.filter().first().days_expire
    expired_dates = []

    if travel_date:
        travel_date = request.POST.get('DateTravel')
    else :
        start_date_str, end_date_str = range_travel.split(' to ')
        
        start_date = datetime.strptime(start_date_str.strip(), '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str.strip(), '%Y-%m-%d')
        
        formatted_dates = []

        current_date = start_date
        while current_date <= end_date:
            formatted_dates.append(current_date.strftime('%d-%m-%Y'))
            current_date += timedelta(days=1)

        formatted_dates_str = ', '.join(formatted_dates)
        travel_date = formatted_dates_str

    duplicate_travel = []
    individual_dates = travel_date.split(',')
    cleaned_dates = ','.join(date.strip() for date in individual_dates)


    if enable_expiry:
        for date_str in individual_dates:
            date_object = datetime.strptime(date_str.strip(), '%d-%m-%Y').date()  # Convert to date
            if (datetime_date.today() - date_object).days >= int(days_expire):
                expired_dates.append(date_object.strftime('%B %d, %Y'))  # Use consistent variable name

        if expired_dates:
            print("Expired dates found:", expired_dates)
            return JsonResponse({'data': 'expired', 'message': ', '.join(expired_dates)})
        else:
            print("All dates are not expired")



    for date in individual_dates:
        cleaned_date = date.strip()

        results = TevIncoming.objects.filter(
            Q(first_name=name) & Q(middle_name=middle) & Q(last_name=lname) &
            Q(date_travel__contains=cleaned_date)
        ).values('date_travel')

        if results:
            duplicate_travel.append(cleaned_date)

    if duplicate_travel:

        formatted_dates = [date.replace("'", "") for date in duplicate_travel]
        result = ",".join(formatted_dates)

        date_components = result.split(',')
        def format_date(date_str):
            date_object = datetime.strptime(date_str, '%d-%m-%Y')
            formatted_date = date_object.strftime('%b. %d, %Y')
            return formatted_date
        formatted_dates = [format_date(date) for date in date_components]
        formatted_dates_string = ', '.join(formatted_dates)
        formatted_dates_string = formatted_dates_string
        
        if date_received:
            tev_add = TevIncoming(
                code=g_code,
                first_name=name,
                middle_name=middle,
                last_name=lname,
                id_no=idd_no,
                account_no=acct_no,
                date_travel=cleaned_dates,
                original_amount=amount,
                incoming_in = date_received,
                remarks=formatted_dates_string,
                user_id=user_id,
                division = division,
                section = section,
                contact_no = contact
            )
        else:
            tev_add = TevIncoming(
                code=g_code,
                first_name=name,
                middle_name=middle,
                last_name=lname,
                id_no=idd_no,
                account_no=acct_no,
                date_travel=cleaned_dates,
                incoming_in = date_time.datetime.now(),
                original_amount=amount,
                remarks=formatted_dates_string,
                user_id=user_id,
                division = division,
                section = section,
                contact_no = contact
            )
        tev_add.save()

        if tev_add.id:
            system_config = SystemConfiguration.objects.first()
            system_config.transaction_code = g_code
            system_config.save()
            last_added_tevincoming = TevIncoming.objects.latest('id')

            for selected_remarks, selected_dates in zip(selected_remarks, selected_dates):
                remarks_lib = Remarks_r(
                    date=selected_dates,
                    incoming_id=last_added_tevincoming.id,
                    remarks_lib_id=selected_remarks
                )
                remarks_lib.save()

        return JsonResponse({'data': 'error', 'message': duplicate_travel})

    else:
        if date_received:
            tev_add = TevIncoming(
                code=g_code,
                first_name=name,
                middle_name=middle,
                last_name=lname,
                id_no=idd_no,
                account_no=acct_no,
                date_travel=cleaned_dates,
                original_amount=amount,
                incoming_in = date_received,
                user_id=user_id,
                division = division,
                section = section,
                contact_no = contact
            )
        else:
            tev_add = TevIncoming(
                code=g_code,
                first_name=name,
                middle_name=middle,
                last_name=lname,
                id_no=idd_no,
                account_no=acct_no,
                date_travel=cleaned_dates,
                original_amount=amount,
                incoming_in = date_time.datetime.now(),
                user_id=user_id,
                division = division,
                section = section,
                contact_no = contact
            )
        tev_add.save()

        last_added_tevincoming = TevIncoming.objects.latest('id')

        for selected_remarks, selected_dates in zip(selected_remarks, selected_dates):
            remarks_lib = Remarks_r(
                date=selected_dates,
                incoming_id=last_added_tevincoming.id,
                remarks_lib_id=selected_remarks
            )
            remarks_lib.save()

        if tev_add.id:
            system_config = SystemConfiguration.objects.first()
            system_config.transaction_code = g_code
            system_config.save()

        return JsonResponse({'data': 'success', 'g_code': g_code})

@csrf_exempt
def out_pending_tev(request):
    user_id = request.session.get('user_id', 0) 
    out_list = request.POST.getlist('out_list[]')
    for item_id  in out_list:
        tev_update = TevIncoming.objects.filter(id=item_id).update(status=2,incoming_out=date_time.datetime.now(), forwarded_by = user_id)
    return JsonResponse({'data': 'success'})


@csrf_exempt
def add_existing_record(request):
    out_list = request.POST.getlist('out_list[]')
    
    for item_id  in out_list:
        tev_update = TevIncoming.objects.filter(id=item_id).update(status=2,incoming_out=date_time.datetime.now())
    
    return JsonResponse({'data': 'success'})


def send_notification(message, contact_number):
    url = 'https://wiserv.dswd.gov.ph/soap/?wsdl'
    try:
        client = Client(url)
        result = client.service.sendMessage(UserName='crgwiservuser', PassWord='#w153rvcr9!', WSID='0',
                                            MobileNo=contact_number, Message=message)
    except Exception:
        pass

def convert_date_string(date_str):
    date_list = date_str.split(',')
    formatted_dates = []
    for date in date_list:
        d, m, y = date.split('-')
        month_name = datetime.strptime(m, '%m').strftime('%B')
        formatted_date = f"{month_name} {d} {y}"
        formatted_dates.append(formatted_date)
    return ', '.join(formatted_dates)



@csrf_exempt
def out_checking_tev(request):

    out_list = request.POST.getlist('out_list[]')  
    user_id = request.session.get('user_id', 0) 
    message = ""

    for item_id in out_list:
        trips_data = TevIncoming.objects.filter(id=item_id).first()  

        if trips_data:

            trips_data.slashed_out = date_time.datetime.now()

            remarks_data = Remarks_r.objects.filter(incoming_id=trips_data.id).values('remarks_lib_id', 'date')
            remarks_list = [
                f"{RemarksLib.objects.get(id=remark['remarks_lib_id']).name} - {remark['date'].strftime('%B %d, %Y')}"
                for remark in remarks_data
            ]
            remarks_str = '; '.join(remarks_list)
            w_remarks_data = remarks_str
            contact_no = trips_data.contact_no
            formatted_dates_list = convert_date_string(trips_data.date_travel).split(', ')
        
            day_format = "%-d" if platform.system() != "Windows" else "%#d"
            formatted_dates_list = [
                datetime.strptime(date, "%B %d %Y").strftime(f"%b. {day_format} %Y")
                for date in formatted_dates_list
            ]

            if len(formatted_dates_list) > 1:
                formatted_dates = f"{formatted_dates_list[0]} to {formatted_dates_list[-1]}"
            else:
                formatted_dates = formatted_dates_list[0]

            

            if trips_data.status_id == 3:  # returned
                if trips_data.remarks:
                    # Ensure correct formatting of remarks
                    formatted_remarks = re.sub(r'(\d{1,2}), (\d{4})', r'\1 \2', trips_data.remarks)
                    formatted_incoming_in = trips_data.incoming_in.strftime("%b. %d %Y")
                    message = "Good day, {}!\n\nYour TE claim for the period of {} was found to be a duplicate of another claim submitted on {} and is subject for a memo\n\n- The DSWD Caraga TRIPS Team.".format(trips_data.first_name.title(), formatted_remarks, formatted_incoming_in)
                    send_notification(message, contact_no)

                elif w_remarks_data and "FORFEITED" not in w_remarks_data:
                    message = "Good day, {}!\n\nYour TE claim for the period of {}, will be returned to your respective division.Please retrieve it for compliance.\n\n- The DSWD Caraga TRIPS Team.".format(trips_data.first_name.title(), formatted_dates, w_remarks_data)
                    send_notification(message, contact_no)

                elif "FORFEITED" in w_remarks_data:
                    message = "Good day, {}!\n\nYour TE claim for the period of {} has been forfeited due to late submission.\n\n- The DSWD Caraga TRIPS Team.".format(trips_data.first_name.title(), formatted_dates)
                    send_notification(message, contact_no)
    
            elif trips_data.status_id == 7 and "FORFEITED" in w_remarks_data: #approved but has forfeited

                message = "Good day, {}!\n\nYour TE claim for the period of {} has been forfeited due to late submission.\n\n- The DSWD Caraga TRIPS Team.".format(trips_data.first_name.title(), formatted_dates)
                send_notification(message, contact_no)
                trips_data.status_id = 4
            else:
                trips_data.status_id = 4  # for payroll
                
            trips_data.review_date_forwarded = date_time.datetime.now()
            trips_data.review_forwarded_by = user_id
            trips_data.save()


    return JsonResponse({'data': 'success'})

@csrf_exempt
def tev_details(request):
    
    tev_id = request.POST.get('tev_id')
    
    result = TevIncoming.objects.filter(id=tev_id).first()
    data = {
        'data': model_to_dict(result)
    }
    return JsonResponse(data)


def review_details(request):
    tev_id = request.POST.get('tev_id')
    with connection.cursor() as cursor:
        query = """
        SELECT
            t1.code,
            t1.first_name,
            t1.middle_name,
            t1.last_name,
            t1.id_no,
            t1.account_no,
            t1.date_travel,
            t1.original_amount,
            t1.final_amount,
            t1.incoming_in,
            t1.incoming_out,
            t1.slashed_out,
            t1.remarks,
            t1.user_id,
            t1.status_id,
            GROUP_CONCAT(t3.id SEPARATOR ', ') AS lacking,
            GROUP_CONCAT(t2.date SEPARATOR ', ') AS date_remarks
        FROM
            tev_incoming t1
            LEFT JOIN remarks_r AS t2 ON t2.incoming_id = t1.id
            LEFT JOIN remarks_lib AS t3 ON t3.id = t2.remarks_lib_id
        WHERE
            t1.id = %s
        """
        cursor.execute(query, [tev_id])
        result = cursor.fetchone()

    # Convert the result to a dictionary for JsonResponse
    if result:
        data = {
            'code': result[0],
            'first_name': result[1],
            'middle_name': result[2],
            'last_name': result[3],
            'id_no': result[4],
            'account_no': result[5],
            'date_travel': result[6],
            'original_amount': result[7],
            'final_amount': result[8],
            'incoming_in': result[9],
            'incoming_out': result[10],
            'slashed_out': result[11],
            'remarks': result[12],
            'user_id': result[13],
            'status_id': result[14],
            'lacking': result[15],
            'date_remarks': result[16],
        }
        return JsonResponse(data)
    else:
        return JsonResponse({'error': 'No data found for the given ID'}, status=404)



@csrf_exempt
def tevemployee(request):
    tev_id = request.POST.get('tev_id')
    qs_object = TevIncoming.objects.filter(id=tev_id).first()
    if qs_object:
        data = serializers.serialize('json', [qs_object])
        return JsonResponse({'data': data})
    else:
        return JsonResponse({'data': None})

@csrf_exempt
def addtev(request):
    
    employeename = request.POST.get('employeename')
    amount = request.POST.get('amount')
    remarks = request.POST.get('remarks')
    user_id = request.session.get('user_id', 0)
    tev_add = TevIncoming(employee_name=employeename,original_amount=amount,incoming_remarks=remarks,user_id=user_id)
    tev_add.save()
    return JsonResponse({'data': 'success'})


@csrf_exempt
def updatetevdetails(request):
    user_id = request.session.get('user_id', 0)
    if request.method == 'POST':
        amount = request.POST.get('final_amount')
        status = request.POST.get('status')
        transaction_id = request.POST.get('transaction_id')
        selected_remarks = request.POST.getlist('selected_remarks[]')
        selected_dates = request.POST.getlist('selected_dates[]')
        TevIncoming.objects.filter(id=transaction_id).update(final_amount=amount,status_id=status, reviewed_by =user_id, date_reviewed =date_time.datetime.now())
        Remarks_r.objects.filter(incoming_id=transaction_id).delete()
        for selected_remarks, selected_dates in zip(selected_remarks, selected_dates):
            Remarks_r.objects.create(
                incoming_id=transaction_id,
                remarks_lib_id=selected_remarks,
                date=selected_dates
        )
        return JsonResponse({'data': 'success'})
    
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})
    
@csrf_exempt
def updatetevamount(request):
    user_id = request.session.get('user_id', 0)
    if request.method == 'POST':
        amount = request.POST.get('final_amount')
        transaction_id = request.POST.get('transaction_id')
        selected_remarks = request.POST.getlist('selected_remarks[]')
        selected_dates = request.POST.getlist('selected_dates[]')
        TevIncoming.objects.filter(id=transaction_id).update(final_amount=amount)
        Remarks_r.objects.filter(incoming_id=transaction_id).delete()
        for selected_remarks, selected_dates in zip(selected_remarks, selected_dates):
            Remarks_r.objects.create(
                incoming_id=transaction_id,
                remarks_lib_id=selected_remarks,
                date=selected_dates
        )
        return JsonResponse({'data': 'success'})
    
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})
    

@csrf_exempt
def delete_entry(request):
    item_id = request.POST.get('item_id')
    user_id = request.session.get('user_id', 0)

    userData = AuthUser.objects.filter(id=user_id).first()
    if not userData:
        return JsonResponse({'error': 'User not found'}, status=404)

    full_name = f"{userData.first_name} {userData.last_name}"

    try:
        with transaction.atomic():
            data = TevIncoming.objects.get(id=item_id)
            dl_codes = TevIncoming.objects.filter(code=data.code)

            for dl_code in dl_codes:
                remark_data = Remarks_r.objects.filter(incoming_id=dl_code.id)
                
                if remark_data: 
                    for remark in remark_data:
                        remarks_lib = RemarksLib.objects.filter(id=remark.remarks_lib_id).first()
                        description = (
                            "This Transaction from RECEIVED module is deleted with code number " + data.code +
                            " and ID Number : " + data.id_no + " Fullname : " + data.first_name + " " +
                            data.middle_name + " " + data.last_name + " with Original amount : " + str(data.original_amount) +
                            " and Date Travel : " + data.date_travel + " deleted by " + full_name +
                            " with remarks " + (remarks_lib.name if remarks_lib else "N/A")
                        )

                        tev_add = TransactionLogs(description=description, user_id=user_id, created_at=timezone.now())
                        tev_add.save()
                        remark.delete()
                else:
                    description = (
                            "This Transaction from RECEIVED module is deleted with code number " + data.code +
                            " and ID Number : " + data.id_no + " Fullname : " + data.first_name + " " +
                            data.middle_name + " " + data.last_name + " with Original amount : " + str(data.original_amount) +
                            " and Date Travel : " + data.date_travel + " deleted by " + full_name
                        )
                    tev_add = TransactionLogs(description=description, user_id=user_id, created_at=timezone.now())
                    tev_add.save()
                dl_code.delete()

            return JsonResponse({'data': 'success'})
    except Exception as e:
        print("Exception occurred:", str(e))
        transaction.rollback()
        return JsonResponse({'error': str(e)}, status=500)
    
@csrf_exempt
def addtevdetails(request):
    return JsonResponse({'data': 'success'})



